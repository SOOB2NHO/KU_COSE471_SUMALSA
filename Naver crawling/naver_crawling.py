import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import time
import os

# 키워드 리스트
keywords = ["안철수"]
start_date_str = "2022.03.02"
end_date_str = "2022.03.03"
max_articles_per_keyword = 1500  # 수집 제한

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
}

# 셀레니움 설정
options = Options()
options.add_argument("--headless")
options.add_argument("--disable-gpu")
options.add_argument("window-size=1920x1080")
driver = webdriver.Chrome(options=options)

def str_to_date(s):
    return datetime.strptime(s, "%Y.%m.%d")

# 기사 본문 파싱 함수 (내용 제외)
def fetch_article_content(url):
    def parse_soup(soup):
        title = (
            soup.select_one(".media_end_head_headline") or
            soup.find("h2") or
            soup.find("div", class_="news_title") or
            soup.find("div", class_="end_tit")
        )
        date = (
            soup.select_one("span.media_end_head_info_datestamp_time._ARTICLE_DATE_TIME") or
            soup.find("em", class_="date") or
            soup.find("span", class_="date") or
            soup.find("div", class_="date")
        )
        if title and date:
            return title.text.strip(), date.text.strip()
        return None, None

    try:
        res = requests.get(url, headers=headers, timeout=5)
        soup = BeautifulSoup(res.text, "html.parser")
        result = parse_soup(soup)
        if all(result):
            return result
    except:
        pass

    try:
        driver.get(url)
        time.sleep(2)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        return parse_soup(soup)
    except:
        return None, None

# ✅ 뉴스 링크 수집 함수
def collect_news_links(search_url):
    try:
        res = requests.get(search_url, headers=headers, timeout=5)
        soup = BeautifulSoup(res.text, "html.parser")
        links = list({a['href'] for a in soup.select(
            "a[href^='https://n.news.naver.com'], a[href^='https://m.entertain.naver.com']") if a.get('href')})
        if len(links) >= 5:
            return links
    except:
        pass
    try:
        driver.get(search_url)
        time.sleep(2)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        links = list({a['href'] for a in soup.select(
            "a[href^='https://n.news.naver.com'], a[href^='https://m.entertain.naver.com']") if a.get('href')})
        return links
    except:
        return []

# ✅ 키워드별 전체 크롤링 루프
for keyword in keywords:
    print(f"\n🔍 키워드 [{keyword}] 크롤링 시작")
    file_name = f"all_news_{keyword}_20250527~20250528.xlsx"

    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "뉴스기사"
        ws.append(['제목', '날짜', 'URL'])

    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 60

    visited_links = set()
    collected = 0
    current_date = str_to_date(start_date_str)
    end_date = str_to_date(end_date_str)

    while current_date <= end_date and collected < max_articles_per_keyword:
        ds = current_date.strftime("%Y.%m.%d")
        de = ds
        print(f"----- 📅 {ds} 크롤링 중 -----")
        fail_count = 0
        max_fails = 5

        for start in range(1, 2000, 10):
            if collected >= max_articles_per_keyword:
                break

            search_url = (
                f"https://search.naver.com/search.naver?where=news&query={keyword}"
                f"&sm=tab_opt&sort=2&photo=0&field=0&pd=3&ds={ds}&de={de}&start={start}"
            )
            news_links = collect_news_links(search_url)

            if not news_links:
                fail_count += 1
                print(f"⚠️ 뉴스 없음 (실패 {fail_count}/{max_fails})")
                if fail_count >= max_fails:
                    print("⛔ 연속 실패 횟수 초과, 종료")
                    break
                continue
            else:
                fail_count = 0

            for href in news_links:
                if href in visited_links or collected >= max_articles_per_keyword:
                    continue
                visited_links.add(href)

                title, date = fetch_article_content(href)
                if title and date:
                    collected += 1
                    print(f"✨ ({collected}) {date} - {title}")
                    ws.append([title, date, href])
                else:
                    print("⚠️ 본문 누락:", href)

            time.sleep(1)

        current_date += timedelta(days=1)

    wb.save(file_name)
    print(f"\n✅ 저장 완료: {file_name}")

driver.quit()